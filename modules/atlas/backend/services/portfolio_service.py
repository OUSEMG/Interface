from datetime import UTC, date, datetime, timedelta
from pathlib import Path
import re
import time

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
DATA_DIR = REPO_ROOT / "data"
LEDGER_PATH = DATA_DIR / "OUSEMG_Holdings_Ledger.xlsx"
SHARES_WORKBOOK_PATH = DATA_DIR / "OUSEMG_Shares_05-01-2026.xlsx"
SHARES_WORKBOOK_PATTERN = "OUSEMG_Shares_*.xlsx"

BENCHMARK_TICKER = "SPY"
CACHE_TTL = 300
PORTFOLIO_SHEETS = {
    "traditional": "Traditional",
    "sustainable": "Sustainable",
}

_cache = {}


def get_cached(key):
    entry = _cache.get(key)
    if entry and (time.time() - entry["ts"]) < CACHE_TTL:
        return entry["data"]
    return None


def set_cached(key, data):
    _cache[key] = {"data": data, "ts": time.time()}


def _to_float(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, str):
        value = value.replace("$", "").replace(",", "").strip()
        if not value:
            return default
    return float(value)


def _to_int(value):
    if value is None or value == "":
        return None
    return int(float(value))


def _is_yes(value):
    return str(value or "").strip().upper() == "YES"


def _resolve_yf_ticker(ticker):
    overrides = {
        "B": "GOLD",
    }
    return overrides.get(ticker, ticker.replace("/", "-"))


def _resolve_holdings_workbook():
    """Use the OUSEMG shares workbook as the portfolio source of truth."""
    if SHARES_WORKBOOK_PATH.exists():
        return SHARES_WORKBOOK_PATH

    candidates = sorted(
        DATA_DIR.glob(SHARES_WORKBOOK_PATTERN),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]

    if LEDGER_PATH.exists():
        return LEDGER_PATH

    raise FileNotFoundError(
        f"Holdings source not found. Expected {SHARES_WORKBOOK_PATH}, "
        f"a {SHARES_WORKBOOK_PATTERN} file, or {LEDGER_PATH} in {DATA_DIR}."
    )


def _date_from_workbook_name(path):
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if not match:
        return None

    month, day, year = match.groups()
    return date(int(year), int(month), int(day))


def _parse_snapshot_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return date(2026, 6, 19)

    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return date(2026, 6, 19)


def _format_snapshot_date(value):
    parsed = _parse_snapshot_date(value)
    return parsed.strftime("%m/%d/%Y")


def read_holdings(sheet_name):
    workbook_path = _resolve_holdings_workbook()
    workbook_snapshot_date = _date_from_workbook_name(workbook_path)
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook sheet '{sheet_name}' was not found")

    ws = wb[sheet_name]
    holdings = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = str(row[1] or "").strip().upper()
        if not ticker:
            break

        if ticker == "CASH":
            holdings.append(
                {
                    "name": row[0] or "Cash",
                    "ticker": "CASH",
                    "yf_ticker": "CASH",
                    "shares": None,
                    "snapshot_price": None,
                    "snapshot_value": _to_float(row[2]),
                    "asset_class": "Cash",
                    "snapshot_date": workbook_snapshot_date,
                    "is_cash": True,
                }
            )
            continue

        shares = _to_int(row[2])
        if not shares:
            continue

        holdings.append(
            {
                "name": row[0],
                "ticker": ticker,
                "yf_ticker": _resolve_yf_ticker(ticker),
                "shares": shares,
                "snapshot_price": None,
                "snapshot_value": 0.0,
                "asset_class": "Equity",
                "snapshot_date": workbook_snapshot_date,
                "is_cash": False,
            }
        )

    return holdings


def _close_frame(raw, tickers):
    if raw.empty or "Close" not in raw:
        return pd.DataFrame(columns=tickers)

    close = raw["Close"]
    if isinstance(close, pd.Series):
        return close.to_frame(tickers[0])
    return close


def _last_price_on_or_before(close, ticker, target_date):
    if ticker not in close:
        return None

    series = close[ticker].dropna()
    if series.empty:
        return None

    filtered = series[series.index.date <= target_date]
    if filtered.empty:
        return None

    return float(filtered.iloc[-1])


def _first_price(close, ticker):
    if ticker not in close:
        return None

    series = close[ticker].dropna()
    if series.empty:
        return None

    return float(series.iloc[0])


def _last_price(close, ticker):
    if ticker not in close:
        return None

    series = close[ticker].dropna()
    if series.empty:
        return None

    return float(series.iloc[-1])


def _previous_price(close, ticker):
    if ticker not in close:
        return None

    series = close[ticker].dropna()
    if len(series) < 2:
        return None

    return float(series.iloc[-2])


def fetch_prices(yf_tickers, snapshot_date):
    # SWAP POINT: Replace this function with BLPAPI equivalent when terminal bridge is ready.
    # BLPAPI call should return the same dict shape:
    # {yf_ticker: {"current": float, "previous": float, "ytd": float, "snapshot": float}}.
    tickers = list(dict.fromkeys([ticker for ticker in yf_tickers if ticker and ticker != "CASH"]))
    all_tickers = list(dict.fromkeys(tickers + [BENCHMARK_TICKER]))
    if not all_tickers:
        return {}

    today = datetime.now(UTC).date()
    start_date = min(date(today.year, 1, 1), snapshot_date) - timedelta(days=10)
    end_date = today + timedelta(days=1)

    raw = yf.download(
        all_tickers,
        start=start_date.strftime("%Y-%m-%d"),
        end=end_date.strftime("%Y-%m-%d"),
        auto_adjust=True,
        progress=False,
    )
    close = _close_frame(raw, all_tickers)

    prices = {}
    for ticker in all_tickers:
        prices[ticker] = {
            "current": _last_price(close, ticker),
            "previous": _previous_price(close, ticker),
            "ytd": _first_price(
                close[close.index.date >= date(today.year, 1, 1)],
                ticker,
            ),
            "snapshot": _last_price_on_or_before(close, ticker, snapshot_date),
        }

    return prices


def _safe_return(current_value, basis_value):
    if basis_value in (None, 0):
        return None
    return (current_value - basis_value) / basis_value


def _position_from_holding(holding, price_data):
    if holding["is_cash"]:
        current_value = holding["snapshot_value"]
        return {
            **holding,
            "current_price": None,
            "previous_price": None,
            "current_value": current_value,
            "market_value": current_value,
            "day_change": 0.0,
            "day_change_percent": 0.0,
            "day_gain_loss": 0.0,
            "gain_loss": 0.0,
            "return_since_snapshot": 0.0,
            "ytd_return": 0.0,
            "weight": 0.0,
            "price_stale": False,
            "_ytd_basis": current_value,
        }

    prices = price_data.get(holding["yf_ticker"], {})
    current_price = prices.get("current")
    previous_price = prices.get("previous")
    ytd_price = prices.get("ytd")
    snapshot_price = prices.get("snapshot")
    price_stale = current_price is None

    if price_stale:
        current_price = snapshot_price
        current_value = current_price * holding["shares"] if current_price else 0.0
    else:
        current_value = current_price * holding["shares"]

    snapshot_value = snapshot_price * holding["shares"] if snapshot_price else 0.0
    day_change = (
        current_price - previous_price
        if current_price is not None and previous_price is not None
        else None
    )
    day_change_percent = _safe_return(current_price, previous_price)
    day_gain_loss = day_change * holding["shares"] if day_change is not None else None
    gain_loss = current_value - snapshot_value if snapshot_value else None
    ytd_basis = ytd_price * holding["shares"] if ytd_price else snapshot_value

    return {
        **holding,
        "snapshot_price": snapshot_price,
        "snapshot_value": snapshot_value,
        "current_price": current_price,
        "previous_price": previous_price,
        "current_value": current_value,
        "market_value": current_value,
        "day_change": day_change,
        "day_change_percent": day_change_percent,
        "day_gain_loss": day_gain_loss,
        "gain_loss": gain_loss,
        "return_since_snapshot": _safe_return(current_value, snapshot_value),
        "ytd_return": _safe_return(current_value, ytd_basis),
        "weight": 0.0,
        "price_stale": price_stale,
        "_ytd_basis": ytd_basis,
    }


def _benchmark_returns(price_data, snapshot_date):
    benchmark = price_data.get(BENCHMARK_TICKER, {})
    current = benchmark.get("current")
    snapshot = benchmark.get("snapshot")
    ytd = benchmark.get("ytd")

    return {
        "benchmark_return_since_snapshot": _safe_return(current, snapshot) if current and snapshot else None,
        "benchmark_ytd_return": _safe_return(current, ytd) if current and ytd else None,
        "snapshot_date": snapshot_date,
    }


def _build_snapshot(portfolio_name, holdings, price_data, fetched_at, cache_hit):
    snapshot_date_value = next((holding["snapshot_date"] for holding in holdings if holding.get("snapshot_date")), None)
    snapshot_date = _parse_snapshot_date(snapshot_date_value)
    positions = [_position_from_holding(holding, price_data) for holding in holdings]

    total_aum = sum(position["current_value"] for position in positions)
    snapshot_aum = sum(position["snapshot_value"] for position in positions)
    ytd_basis = sum(position["_ytd_basis"] for position in positions)

    for position in positions:
        position["weight"] = position["current_value"] / total_aum if total_aum else 0.0
        position["snapshot_date"] = _format_snapshot_date(position["snapshot_date"])
        position.pop("_ytd_basis", None)

    benchmark = _benchmark_returns(price_data, snapshot_date)
    return_since_snapshot = _safe_return(total_aum, snapshot_aum)
    ytd_return = _safe_return(total_aum, ytd_basis)
    benchmark_since_snapshot = benchmark["benchmark_return_since_snapshot"]
    benchmark_ytd = benchmark["benchmark_ytd_return"]

    return {
        "portfolio": portfolio_name,
        "fetched_at": fetched_at,
        "cache_hit": cache_hit,
        "snapshot_date": _format_snapshot_date(snapshot_date),
        "benchmark_ticker": BENCHMARK_TICKER,
        "summary": {
            "total_aum": total_aum,
            "snapshot_aum": snapshot_aum,
            "total_gain_loss": total_aum - snapshot_aum,
            "return_since_snapshot": return_since_snapshot,
            "ytd_return": ytd_return,
            "benchmark_return_since_snapshot": benchmark_since_snapshot,
            "benchmark_ytd_return": benchmark_ytd,
            "excess_return_since_snapshot": (
                return_since_snapshot - benchmark_since_snapshot
                if return_since_snapshot is not None and benchmark_since_snapshot is not None
                else None
            ),
            "excess_ytd_return": (
                ytd_return - benchmark_ytd
                if ytd_return is not None and benchmark_ytd is not None
                else None
            ),
        },
        "holdings": positions,
    }


def get_portfolio_snapshot(portfolio_name, force_refresh=False):
    normalized = portfolio_name.strip().lower()
    if normalized not in PORTFOLIO_SHEETS:
        raise ValueError("portfolio_name must be either 'traditional' or 'sustainable'")

    cache_key = f"portfolio:{normalized}"
    if not force_refresh:
        cached = get_cached(cache_key)
        if cached:
            return {**cached, "cache_hit": True}

    holdings = read_holdings(PORTFOLIO_SHEETS[normalized])
    snapshot_date_value = next((holding["snapshot_date"] for holding in holdings if holding.get("snapshot_date")), None)
    snapshot_date = _parse_snapshot_date(snapshot_date_value)
    price_data = fetch_prices([holding["yf_ticker"] for holding in holdings], snapshot_date)
    fetched_at = datetime.now(UTC).isoformat().replace("+00:00", "Z")

    snapshot = _build_snapshot(normalized, holdings, price_data, fetched_at, False)
    set_cached(cache_key, snapshot)
    return snapshot


def _weighted_average_summary(left, right, field):
    total_snapshot = left["summary"]["snapshot_aum"] + right["summary"]["snapshot_aum"]
    if not total_snapshot:
        return None

    left_value = left["summary"].get(field)
    right_value = right["summary"].get(field)
    if left_value is None or right_value is None:
        return None

    return (
        (left_value * left["summary"]["snapshot_aum"])
        + (right_value * right["summary"]["snapshot_aum"])
    ) / total_snapshot


def get_combined_snapshot(force_refresh=False):
    traditional = get_portfolio_snapshot("traditional", force_refresh)
    sustainable = get_portfolio_snapshot("sustainable", force_refresh)

    holdings = [
        {**holding, "portfolio": "traditional"}
        for holding in traditional["holdings"]
    ] + [
        {**holding, "portfolio": "sustainable"}
        for holding in sustainable["holdings"]
    ]

    total_aum = traditional["summary"]["total_aum"] + sustainable["summary"]["total_aum"]
    snapshot_aum = traditional["summary"]["snapshot_aum"] + sustainable["summary"]["snapshot_aum"]
    for holding in holdings:
        holding["weight"] = holding["current_value"] / total_aum if total_aum else 0.0

    benchmark_since_snapshot = traditional["summary"]["benchmark_return_since_snapshot"]
    benchmark_ytd = traditional["summary"]["benchmark_ytd_return"]
    return_since_snapshot = _safe_return(total_aum, snapshot_aum)
    ytd_return = _weighted_average_summary(traditional, sustainable, "ytd_return")

    return {
        "portfolio": "combined",
        "combined_aum": total_aum,
        "fetched_at": max(traditional["fetched_at"], sustainable["fetched_at"]),
        "cache_hit": traditional["cache_hit"] and sustainable["cache_hit"],
        "snapshot_date": traditional["snapshot_date"],
        "benchmark_ticker": BENCHMARK_TICKER,
        "summary": {
            "total_aum": total_aum,
            "snapshot_aum": snapshot_aum,
            "total_gain_loss": total_aum - snapshot_aum,
            "return_since_snapshot": return_since_snapshot,
            "ytd_return": ytd_return,
            "benchmark_return_since_snapshot": benchmark_since_snapshot,
            "benchmark_ytd_return": benchmark_ytd,
            "excess_return_since_snapshot": (
                return_since_snapshot - benchmark_since_snapshot
                if return_since_snapshot is not None and benchmark_since_snapshot is not None
                else None
            ),
            "excess_ytd_return": (
                ytd_return - benchmark_ytd
                if ytd_return is not None and benchmark_ytd is not None
                else None
            ),
        },
        "traditional": traditional,
        "sustainable": sustainable,
        "holdings": holdings,
    }


def get_top_movers(portfolio_name):
    snapshot = get_portfolio_snapshot(portfolio_name)
    ranked = [
        holding
        for holding in snapshot["holdings"]
        if not holding["is_cash"] and holding["return_since_snapshot"] is not None
    ]
    ranked.sort(key=lambda holding: holding["return_since_snapshot"], reverse=True)

    bottom = sorted(ranked, key=lambda holding: holding["return_since_snapshot"])
    return {
        "top": ranked[:5],
        "bottom": bottom[:5],
    }
